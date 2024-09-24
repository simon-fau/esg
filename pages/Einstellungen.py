import streamlit as st
import os
import shutil
from openpyxl import load_workbook
import pandas as pd
import altair as alt
from vl_convert import vl_convert as altair_save
from openpyxl.drawing.image import Image as ExcelImage

# Define the file paths
STATE_FILE = 'SessionStates.pkl'
BACKUP_STATE_FILE = 'ab.pkl'
DEFAULT_STATE_FILE = 'Grundlagen.pkl'
DEFAULT_BACKUP_FILE = 'Grundlagen_Themen_ESRS.pkl'

# Function to reset the session state
def reset_session_state():
    st.session_state.clear()  # Clear all session state values

    # Check if 'SessionStates.pkl' and 'ab.pkl' exist, and overwrite them with the default files
    if os.path.exists(STATE_FILE) and os.path.exists(BACKUP_STATE_FILE):
        shutil.copy(DEFAULT_STATE_FILE, STATE_FILE)  # Overwrite 'SessionStates.pkl' with 'Grundlagen.pkl'
        shutil.copy(DEFAULT_BACKUP_FILE, BACKUP_STATE_FILE)  # Overwrite 'ab.pkl' with 'Grundlagen_Themen_ESRS.pkl'
        st.success("Session state has been reset, and the default settings have been restored for both 'SessionStates.pkl' and 'ab.pkl'.")
    else:
        st.error("State files not found. Please ensure 'SessionStates.pkl', 'ab.pkl', 'Grundlagen.pkl', and 'Grundlagen_Themen_ESRS.pkl' exist in the correct directory.")

# Function to simulate the modal dialog
def show_modal_dialog():
    st.session_state.modal_open = True  # Flag to open the modal
    with st.expander("⚠️ Bestätigen Sie Ihre Aktion", expanded=True):
        st.write("Sind Sie sicher, dass Sie alle gespeicherten Inhalte aus der App entfernen möchten?")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Ja"):
                reset_session_state()
                st.session_state.modal_open = False
        with col2:
            if st.button("Nein"):
                st.session_state.modal_open = False
                st.warning("Zurücksetzung abgebrochen.")

# Display the settings page
def display_settings_page():
    st.header("Einstellungen")
    st.markdown("Auf dieser Seite können Sie Ihre Einstellungen verwalten.")

    if 'modal_open' not in st.session_state:
        st.session_state.modal_open = False

    if st.button('🔄 App neu starten'):
        st.session_state.modal_open = True  # Trigger modal on button press
    
    if st.session_state.modal_open:
        show_modal_dialog()  # Show the modal dialog if triggered

# Function to create the Excel file based on filtered_df and Allgemeine_Angaben from session_state
def Ausleitung_Excel():
    # Check if the filtered_df exists in the session state
    if 'filtered_df' not in st.session_state:
        st.error("Es gibt keine gefilterten Daten in der Session-State.")
        return

    # Load the existing 'Ergebnisse.xlsx' file
    template_file = 'Ergebnisse.xlsx'
    
    if not os.path.exists(template_file):
        st.error(f"Die Datei {template_file} wurde nicht gefunden.")
        return

    # Create a new workbook based on the template
    try:
        workbook = load_workbook(template_file)
    except Exception as e:
        st.error(f"Fehler beim Laden der Excel-Datei: {str(e)}")
        return
    
    # Check if the data required for the chart is available in session_state
    if 'selected_columns' in st.session_state and len(st.session_state['selected_columns']) > 0:
        selected_columns = st.session_state['selected_columns']
        
        # Prepare the data
        selected_columns_df = pd.DataFrame(selected_columns)
        columns_to_display = ['Score Finanzen', 'Score Auswirkung', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Stakeholder Wichtigkeit']
        
        # Check if Thema column exists and handle missing values
        if 'Thema' not in selected_columns_df.columns:
            st.error("Die Spalte 'Thema' fehlt in den Daten.")
            return
        
        selected_columns_df = selected_columns_df.dropna(subset=['Thema'])
        selected_columns_df['Thema'] = selected_columns_df['Thema'].astype(str)

        def assign_color(theme):
            if theme in ['Klimawandel', 'Umweltverschmutzung', 'Wasser- & Meeresressourcen', 'Biodiversität', 'Kreislaufwirtschaft']:
                return 'Environmental'
            elif theme in ['Eigene Belegschaft', 'Belegschaft Lieferkette', 'Betroffene Gemeinschaften', 'Verbraucher und Endnutzer']:
                return 'Social'
            elif theme == 'Unternehmenspolitik':
                return 'Governance'
            else:
                return 'Sonstige'
        
        selected_columns_df['color'] = selected_columns_df['Thema'].apply(assign_color)

        # Create the base scatter chart
        scatter = alt.Chart(selected_columns_df, width=1000, height=800).mark_circle().encode(
            x=alt.X('Score Finanzen', scale=alt.Scale(domain=(0, 1000)), title='Finanzielle Wesentlichkeit'),
            y=alt.Y('Score Auswirkung', scale=alt.Scale(domain=(0, 1000)), title='Auswirkungsbezogene Wesentlichkeit'),
            color=alt.Color('color:N', scale=alt.Scale(
                domain=['Environmental', 'Social', 'Governance', 'Sonstige'],
                range=['green', 'yellow', 'blue', 'gray']
            ), legend=alt.Legend(
                title="Thema",
                orient="right",
                titleColor='black',
                labelColor='black',
                titleFontSize=12,
                labelFontSize=10,
                values=['Environmental', 'Social', 'Governance', 'Sonstige']
            )),
            size=alt.Size('Stakeholder Wichtigkeit:Q', scale=alt.Scale(range=[100, 1000]), legend=alt.Legend(
                title="Stakeholder Wichtigkeit",
                orient="right",
                titleColor='black',
                labelColor='black',
                titleFontSize=12,
                labelFontSize=10
            )),
            tooltip=['Score Finanzen', 'Score Auswirkung', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Stakeholder Wichtigkeit']
        )

        # Add red diagonal line
        line = alt.Chart(pd.DataFrame({
            'x': [0, st.session_state['intersection_value']],
            'y': [st.session_state['intersection_value'], 0]
        })).mark_line(color='red').encode(
            x='x:Q',
            y='y:Q'
        )

        # Add shaded area under the line
        area = alt.Chart(pd.DataFrame({
            'x': [0, st.session_state['intersection_value']],
            'y': [st.session_state['intersection_value'], 0]
        })).mark_area(opacity=0.3, color='lightcoral').encode(
            x='x:Q',
            y='y:Q'
        )

        # Combine scatter, line, and area
        chart = area + scatter + line

        # Save the chart as a PNG file using vl-convert
        chart_png_file = 'scatter_chart.png'
        chart.save(chart_png_file, format='png')
        
        chart_sheet = workbook['Shortlist']
        chart_sheet['G1'] = "Schwellenwert für wesentliche Themen:"
        chart_sheet['G2'] = "Schwellenwert für Stakeholder Wichtigkeit:"
        chart_sheet['G3'] = "Wesentlichkeitsmatrix der Shortlist:"

        # Insert the PNG image into the Excel sheet
        img = ExcelImage(chart_png_file)
        img.width, img.height = 600, 400  # Set desired size of the image
        chart_sheet.add_image(img, 'G5')  # Place the image in cell B2
        chart_sheet['K1'] = st.session_state['intersection_value']  # Insert the intersection value
        chart_sheet['K2'] = st.session_state['stakeholder_importance_value']

    else:
        st.info("Keine Daten für die Grafik vorhanden.")

    #----------Shortlist Sheet----------#

    # Ensure that the 'Shortlist' sheet exists
    if 'Shortlist' not in workbook.sheetnames:
        st.error("Das Blatt 'Shortlist' existiert nicht in der Datei.")
        return

    # Select the 'Shortlist' worksheet
    shortlist_sheet = workbook['Shortlist']

    # Retrieve the filtered data from session_state
    dataframe = st.session_state.filtered_df

    # Write the filtered data (Thema, Unterthema, Unter-Unterthema) into the 'Shortlist' sheet
    first_empty_row = 1
    while shortlist_sheet[f'A{first_empty_row}'].value or shortlist_sheet[f'B{first_empty_row}'].value or shortlist_sheet[f'C{first_empty_row}'].value:
        first_empty_row += 1

    # Schreibe die gefilterten Daten (Thema, Unterthema, Unter-Unterthema) in das 'Shortlist'-Blatt
    for index, row in dataframe.iterrows():
        shortlist_sheet[f'A{first_empty_row}'] = row['Thema']
        shortlist_sheet[f'B{first_empty_row}'] = row['Unterthema']
        shortlist_sheet[f'C{first_empty_row}'] = row['Unter-Unterthema']
        shortlist_sheet[f'D{first_empty_row}'] = row['Score Finanzen']
        shortlist_sheet[f'E{first_empty_row}'] = row['Score Auswirkung']
        first_empty_row += 1

    #----------Allgemeine Angaben Sheet----------#

    # Check if Antwort exists in session state
    if 'Antwort' in st.session_state:
        # Ensure that the 'Allgemeine Angaben' sheet exists
        if 'Allgemeine Angaben' not in workbook.sheetnames:
            st.error("Das Blatt 'Allgemeine Angaben' existiert nicht in der Datei.")
            return

        # Select the 'Allgemeine Angaben' worksheet
        general_sheet = workbook['Allgemeine Angaben']

        # Write Antwort data into column C starting at row 2
        antwort_data = st.session_state.Antwort
        row_index = 2  # Start from row 2

        for entry in antwort_data:
            general_sheet[f'C{row_index}'] = entry
            row_index += 1

    #----------Mindestangaben Sheet----------#

    # Add Mindestangaben to the "Mindestangaben" sheet if it exists
    if 'Antworten_MDR' in st.session_state:
        if 'Mindestangaben' not in workbook.sheetnames:
            st.error("Das Blatt 'Mindestangaben' existiert nicht in der Datei.")
            return
        
        # Select the 'Mindestangaben' worksheet
        min_sheet = workbook['Mindestangaben']

        # Write the Antworten_MDR data into column C starting at row 2
        mdr_data = st.session_state['Antworten_MDR']
        row_index = 2  # Start from row 2 in the Mindestangaben sheet

        for entry in mdr_data:
            min_sheet[f'C{row_index}'] = entry
            row_index += 1

    #----------Stakeholder Ranking----------#

    # Check if the 'Stakeholder' sheet exists
    if 'Stakeholder' not in workbook.sheetnames:
        st.error("Das Blatt 'Stakeholder' existiert nicht in der Datei.")
        return
    
    # Select the 'Stakeholder' worksheet
    stakeholder_sheet = workbook['Stakeholder']
    
    # Get the ranking DataFrame from session_state
    if 'ranking_table' not in st.session_state:
        st.error("Es gibt keine ranking_table in der Session-State.")
        return
    
    ranking_table = st.session_state['ranking_table']

    # Write the ranking_table DataFrame to 'Stakeholder' sheet starting at row 2
    first_empty_row_stakeholder = 3 # Start from the first empty row

    for index, row in ranking_table.iterrows():
        stakeholder_sheet[f'A{first_empty_row_stakeholder}'] = row['Ranking']
        stakeholder_sheet[f'B{first_empty_row_stakeholder}'] = row['Gruppe']
        stakeholder_sheet[f'C{first_empty_row_stakeholder}'] = row['Score']
        first_empty_row_stakeholder += 1


    # Export the 'table_right_df' DataFrame to 'STakeholder' starting at cell E2
    if 'Einbezogene_Stakeholder' in st.session_state and st.session_state.Einbezogene_Stakeholder:
        # Filter the ranking_table for included stakeholders
        table_right_df = st.session_state['ranking_table'][st.session_state['ranking_table']['Gruppe'].isin(st.session_state.Einbezogene_Stakeholder)]

        # Start writing to column E, row 2
        row = 3
        for index, data in table_right_df.iterrows():
            stakeholder_sheet[f'E{row}'] = data['Ranking']
            stakeholder_sheet[f'F{row}'] = data['Gruppe']
            stakeholder_sheet[f'G{row}'] = data['Score']
            row += 1

    #----------Externe Nachhaltigkeitspunkte----------#

    # Export the 'stakeholder_punkte_filtered' DataFrame to 'Externe Nachhaltigkeitspunkte' starting at row 2
    if 'stakeholder_punkte_filtered' in st.session_state:
        # Ensure that the 'Externe Nachhaltigkeitspunkte' sheet exists
        if 'Externe Nachhaltigkeitspunkte' not in workbook.sheetnames:
            # Create a new sheet if it doesn't exist
            workbook.create_sheet('Externe Nachhaltigkeitspunkte')

        # Select the 'Externe Nachhaltigkeitspunkte' worksheet
        sustainability_sheet = workbook['Externe Nachhaltigkeitspunkte']

        # Write the DataFrame to the sheet starting at row 2
        punkt_df = st.session_state['stakeholder_punkte_filtered']

        # Write specific columns to the Excel sheet
        row_start = 2
        for idx, row in punkt_df.iterrows():
            sustainability_sheet[f'A{row_start}'] = row.get('Platzierung', '')  # Platzierung in column A
            sustainability_sheet[f'B{row_start}'] = row.get('Thema', '')        # Thema in column B
            sustainability_sheet[f'C{row_start}'] = row.get('Unterthema', '')   # Unterthema in column C
            sustainability_sheet[f'D{row_start}'] = row.get('Unter-Unterthema', '') # Unter-Unterthema in column D
            sustainability_sheet[f'E{row_start}'] = row.get('Stakeholder Bew Auswirkung', '')       
            sustainability_sheet[f'F{row_start}'] = row.get('Stakeholder Bew Finanzen', '')
            sustainability_sheet[f'G{row_start}'] = row.get('Stakeholder Gesamtbew', '')
            sustainability_sheet[f'H{row_start}'] = row.get('Stakeholder', '')
            sustainability_sheet[f'I{row_start}'] = row.get('Quelle', '')       # Quelle in column J
            row_start += 1

    #----------Longlist----------#

    # Ensure that the 'Longlist' sheet exists
    if 'Longlist' not in workbook.sheetnames:
        # Create a new sheet if it doesn't exist
        workbook.create_sheet('Longlist')

    # Select the 'Longlist' worksheet
    longlist_sheet = workbook['Longlist']

    # Check if longlist exists in the session state
    if 'longlist' in st.session_state:
        # Retrieve longlist from session state
        longlist_df = st.session_state['longlist']

        for r_idx, row in longlist_df.iterrows():
            for c_idx, value in enumerate(row, start=1):
                longlist_sheet.cell(row=r_idx + 2, column=c_idx, value=value)  # Write data starting from row 2
    else:
        st.info("Keine Daten für 'Longlist' vorhanden.")

       #----------Interne Nachhaltigkeitspunkte----------#

    # Ensure that the 'Interne Nachhaltigkeitspunkte' sheet exists
    if 'Interne Nachhaltigkeitspunkte' not in workbook.sheetnames:
        # Create a new sheet if it doesn't exist
        workbook.create_sheet('Interne Nachhaltigkeitspunkte')

    # Select the 'Interne Nachhaltigkeitspunkte' worksheet
    internal_sustainability_sheet = workbook['Interne Nachhaltigkeitspunkte']

    # Check if df2 exists in the session state
    if 'df2' in st.session_state:
        # Retrieve df2 from session state
        df2 = st.session_state['df2']

        # Write the data to the 'Interne Nachhaltigkeitspunkte' sheet
        for r_idx, row in df2.iterrows():
            for c_idx, value in enumerate(row, start=1):
                internal_sustainability_sheet.cell(row=r_idx + 2, column=c_idx, value=value)  # Write data starting from row 3
    else:
        st.info("Keine Daten für 'Interne Nachhaltigkeitspunkte' vorhanden.")

    # Save the new Excel file with the name 'Ergebnisse_WA.xlsx'
    output_file = 'Ergebnisse_WA.xlsx'
    
    try:
        workbook.save(output_file)
        st.success(f"Die Datei '{output_file}' wurde erfolgreich erstellt.")
        
        # Provide a download link
        with open(output_file, 'rb') as file:
            btn = st.download_button(
                label="Ergebnisse_WA.xlsx herunterladen",
                data=file,
                file_name=output_file,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    except Exception as e:
        st.error(f"Fehler beim Speichern der Excel-Datei: {str(e)}")


def display_page():
    col1, colplatzhalter, col2 = st.columns([1, 1, 1])
    with col1:
        display_settings_page()

    with colplatzhalter:
        st.write("")
    
    with col2:
        Ausleitung_Excel()

    






