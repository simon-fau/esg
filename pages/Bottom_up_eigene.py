import streamlit as st
import pandas as pd
import pickle
import os
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode
from openpyxl import load_workbook
import shutil
import io

# Datei zum Speichern des Sitzungszustands
state_file = 'session_statekizz.pkl'
template_path = r'C:\Users\andre\OneDrive\Dokumente\Stakeholder_Input_Vorlage_V1.xlsx'

# Funktion zum Laden des Sitzungszustands
def load_session_state():
    if os.path.exists(state_file):
        with open(state_file, 'rb') as f:
            return pickle.load(f)
    else:
        return {}

# Funktion zum Speichern des Sitzungszustands
def save_session_state(state):
    # Laden des aktuellen Zustands
    current_state = load_session_state()
    # Hinzufügen des neuen Zustands zum aktuellen Zustand
    combined_state = {**current_state, **state}
    # Speichern des kombinierten Zustands
    with open(state_file, 'wb') as f:
        pickle.dump(combined_state, f)
    
# Laden des Sitzungszustands aus der Datei
loaded_state = load_session_state()
# Überprüfen, ob 'df2' im Sitzungszustand vorhanden ist, und initialisieren Sie ihn andernfalls ohne Zeilen
if 'df2' not in st.session_state:
    st.session_state.df2 = pd.DataFrame(columns=["Thema", "Unterthema", "Unter-Unterthema"])
if 'df2' in loaded_state:
    st.session_state.df2 = loaded_state['df2']

def eigene_punkte():
    with st.sidebar:
        st.markdown("---")
        st.write("**Inhalte hinzufügen**")
        thema = st.selectbox('Thema auswählen',
        options=[
            'Klimawandel', 
            'Umweltverschmutzung', 
            'Wasser- und Meeresressourcen', 
            'Biologische Vielfalt und Ökosysteme', 
            'Kreislaufwirtschaft',
            'Eigene Belegschaft',
            'Arbeitskräfte in der Wertschöpfungskette',
            'Betroffene Gemeinschaften',
            'Verbraucher und End-nutzer',
            'Unternehmenspolitik'
            ], 
        index=0, 
        key='thema'
    )
    
        if thema == 'Klimawandel':
            unterthema_options = ['Anpassung an den Klimawandel', 'Klimaschutz', 'Energie']
        elif thema == 'Umweltverschmutzung':
            unterthema_options = [
                'Luftverschmutzung', 
                'Wasserverschmutzung', 
                'Bodenverschmutzung', 
                'Verschmutzung von lebenden Organismen und Nahrungsressourcen', 
                'Besorgniserregende Stoffe', 
                'Mikroplastik'
            ]
        elif thema == 'Wasser- und Meeresressourcen':
            unterthema_options = ['Wasser', 'Meeresressourcen']
        elif thema == 'Biologische Vielfalt und Ökosysteme':
            unterthema_options = [
                'Direkte Ursachen des Biodiversitätsverlusts', 
                'Auswirkungen auf den Zustand der Arten', 
                'Auswirkungen auf den Umfang und den Zustand von Ökosystemen'
            ]
        elif thema == 'Kreislaufwirtschaft':
            unterthema_options = [
                'Auswirkungen und Abhängigkeiten von Ökosystemdienstleistungen', 
                'Ressourcenzuflüsse, einschließlich Ressourcennutzung', 
                'Ressourcenabflüsse im Zusammenhang mit Produkten und Dienstleistungen', 
                'Abfälle'
            ]
        elif thema == 'Eigene Belegschaft':
            unterthema_options = [
                'Arbeitsbedingungen',
                'Gleichbehandlung und Chancengleichheit für alle',
                'Sonstige arbeitsbezogene Rechte'
            ]
        elif thema == 'Arbeitskräfte in der Wertschöpfungskette':
            unterthema_options = [
                'Arbeitsbedingungen',
                'Gleichbehandlung und Chancengleichheit für alle',
                'Sonstige arbeitsbezogene Rechte'
            ]
        elif thema == 'Betroffene Gemeinschaften':
            unterthema_options = [
                'Wirtschaftliche, soziale und kulturelle Rechte von Gemeinschaften',
                'Bürgerrechte und politische Rechte von Gemeinschaften',
                'Rechte indigener Völker'
            ]
        elif thema == 'Verbraucher und End-nutzer':
            unterthema_options = [
                'Informationsbezogene Auswirkungen für Verbraucher und/oder Endnutzer',
                'Persönliche Sicherheit von Verbrauchern und/oder Endnutzern',
                'Soziale Inklusion von Verbrauchern und/oder Endnutzern'
            ]

        elif thema == 'Unternehmenspolitik':
            unterthema_options = [
                'Unternehmenskultur',
                'Schutz von Hinweisgebern (Whistleblowers)',
                'Tierschutz',
                'Politisches Engagement und Lobbytätigkeiten',
                'Management der Beziehungen zu Lieferanten, einschließlich Zahlungspraktiken',
                'Korruption und Bestechung'
            ]

        unterthema = st.selectbox('Unterthema auswählen', options=unterthema_options, index=0, key='unterthema')
        unter_unterthema = st.text_input('Unter-Unterthema eingeben', key='unter_unterthema')    
        add_row = st.button('Hinzufügen', key='add_row')

        if add_row:
            empty_row_index = st.session_state.df2[(st.session_state.df2["Thema"] == "") & (st.session_state.df2["Unterthema"] == "") & (st.session_state.df2["Unter-Unterthema"] == "")].first_valid_index()
            if empty_row_index is not None:
                st.session_state.df2.at[empty_row_index, "Thema"] = thema
                st.session_state.df2.at[empty_row_index, "Unterthema"] = unterthema
                st.session_state.df2.at[empty_row_index, "Unter-Unterthema"] = unter_unterthema
            else:
                new_row = {"Thema": thema, "Unterthema": unterthema, "Unter-Unterthema": unter_unterthema}
                st.session_state.df2 = st.session_state.df2._append(new_row, ignore_index=True)
            save_session_state({'df2': st.session_state.df2})

    gb = GridOptionsBuilder.from_dataframe(st.session_state.df2)
    gb.configure_default_column(editable=True, resizable=True, sortable=True, filterable=True)
    gb.configure_grid_options(domLayout='autoHeight', enableRowId=True, rowId='index')
    grid_options = gb.build()
    grid_options['columnDefs'] = [{'checkboxSelection': True, 'headerCheckboxSelection': True, 'width': 50}] + grid_options['columnDefs']

    if st.session_state.df2.empty:
        st.warning("Keine Daten vorhanden. Bitte fügen Sie über die Sidebar Inhalte hinzu.")

    grid_response = AgGrid(
        st.session_state.df2.reset_index(),
        gridOptions=grid_options,
        fit_columns_on_grid_load=True,
        height=300,
        width='100%',
        update_mode=GridUpdateMode.MODEL_CHANGED,
        allow_unsafe_jscode=True,
        return_mode=DataReturnMode.__members__['AS_INPUT'],  # Adjust the DataReturnMode as per available options
        selection_mode='multiple'
    )

    st.sidebar.markdown("---")
    st.sidebar.write("**Inhalte bearbeiten**")
    add_empty_row = st.sidebar.button('Leere Zeile hinzufügen', key='add_empty_row')
    if add_empty_row:
        empty_row = {"Thema": "", "Unterthema": "", "Unter-Unterthema": ""}
        st.session_state.df2 = st.session_state.df2._append(empty_row, ignore_index=True)
        save_session_state({'df2': st.session_state.df2})
        st.experimental_rerun()

    delete_rows = st.sidebar.button('Ausgewählte Zeilen löschen', key='delete_rows')
    if delete_rows:
        selected_rows = grid_response['selected_rows']
        selected_indices = [row['index'] for row in selected_rows]
        st.session_state.df2 = st.session_state.df2.drop(selected_indices)
        save_session_state({'df2': st.session_state.df2})
        st.experimental_rerun()

    save_changes = st.button('Änderungen speichern', key='save_changes')
    if save_changes:
        st.session_state.df2 = grid_response['data'].set_index('index')
        save_session_state({'df2': st.session_state.df2})
    
    # Informationsnachricht unter dem Button
    st.caption("ℹ️ Sie müssen diesen Button nur drücken, wenn Sie Inhalte direkt in die Tabelle geschrieben haben. Achten Sie darauf, dass Sie die Inhalte mit Enter bestätigen.")

    # Button zum Übertragen der Inhalte in die Excel-Datei
    if st.button('Inhalte zur Excel hinzufügen'):
        transfer_data_to_excel(st.session_state.df2)


def transfer_data_to_excel(dataframe):
    # Kopie der Template-Datei erstellen
    temp_excel_path = 'Stakeholder_Input_Vorlage_V1_Copy.xlsx'
    shutil.copyfile(template_path, temp_excel_path)

    # Laden der Kopie der Excel-Datei
    workbook = load_workbook(temp_excel_path)
    sheet = workbook['Eigene']

    first_empty_row = 2

    # Übertragen der Daten in die Excel-Datei
    for index, row in dataframe.iterrows():
        sheet[f'A{first_empty_row}'] = "Eigene"
        sheet[f'B{first_empty_row}'] = row['Thema']
        sheet[f'C{first_empty_row}'] = row['Unterthema']
        sheet[f'D{first_empty_row}'] = row['Unter-Unterthema']
        first_empty_row += 1

    # Speichern der bearbeiteten Kopie der Excel-Datei
    workbook.save(temp_excel_path)
    st.success('Inhalte erfolgreich zur Excel-Datei hinzugefügt.')

def download_excel():
    # Pfad zur kopierten und bearbeiteten Excel-Datei
    temp_excel_path = 'Stakeholder_Input_Vorlage_V1_Copy.xlsx'
    workbook = load_workbook(temp_excel_path)
    with io.BytesIO() as virtual_workbook:
        workbook.save(virtual_workbook)
        virtual_workbook.seek(0)
        return virtual_workbook.read()

def display_page():
    eigene_punkte()
    # Download-Button für die Excel-Datei
    if st.download_button(label="Excel-Datei herunterladen",
                          data=download_excel(),
                          file_name="Stakeholder_Input.xlsx",
                          mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
        st.success("Download gestartet!")







