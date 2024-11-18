import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import altair as alt
from vl_convert import vl_convert as altair_save
from openpyxl.drawing.image import Image as ExcelImage
import os
from pages.Themenspezifische_ESRS import calculate_percentages
from pages.Longlist import bewertung_Uebersicht_Nein

# Funktion zur Erstellung einer Excel-Datei basierend auf dem gefilterten DataFrame und den allgemeinen Angaben aus dem Session-State
def Ausleitung_Excel():
    # Überprüfung, ob 'filtered_df' im Session-State vorhanden ist
    if 'filtered_df' not in st.session_state:
        st.error("Es gibt keine gefilterten Daten in der Session-State.")
        return

    # Laden der vorhandenen 'Ergebnisse.xlsx'-Datei als Vorlage
    template_file = os.path.join(os.path.dirname(__file__), 'Templates', 'Ergebnisse.xlsx')
    
    # Überprüfung, ob die Vorlagendatei existiert
    if not os.path.exists(template_file):
        st.error(f"Die Datei {template_file} wurde nicht gefunden.")
        return

    # Erstellen einer neuen Arbeitsmappe basierend auf der Vorlage
    try:
        workbook = load_workbook(template_file)
    except Exception as e:
        st.error(f"Fehler beim Laden der Excel-Datei: {str(e)}")
        return
    
    #---------- Wesentlichkeitsmatrix der Shortlist ----------#

    # Überprüfung, ob die für das Diagramm benötigten Daten im Session-State verfügbar sind
    if 'selected_columns' in st.session_state and len(st.session_state['selected_columns']) > 0:
        selected_columns = st.session_state['selected_columns']
        
        # Daten vorbereiten
        selected_columns_df = pd.DataFrame(selected_columns)
        # Spalten, die im Diagramm angezeigt werden sollen
        columns_to_display = ['Score Finanzen', 'Score Auswirkung', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Stakeholder Wichtigkeit']
        
        # Überprüfung, ob die Spalte 'Thema' existiert, und Umgang mit fehlenden Werten
        if 'Thema' not in selected_columns_df.columns:
            st.error("Die Spalte 'Thema' fehlt in den Daten.")
            return
        
        # Entfernen von Zeilen mit fehlenden Werten in der Spalte 'Thema'
        selected_columns_df = selected_columns_df.dropna(subset=['Thema'])
        selected_columns_df['Thema'] = selected_columns_df['Thema'].astype(str)

        # Funktion zur Zuweisung einer Farbe basierend auf dem Thema
        def assign_color(theme):
            if theme in ['Klimawandel', 'Umweltverschmutzung', 'Wasser- & Meeresressourcen', 'Biodiversität', 'Kreislaufwirtschaft']:
                return 'Environmental'  # Zuordnung der Kategorie "Umwelt"
            elif theme in ['Eigene Belegschaft', 'Belegschaft Lieferkette', 'Betroffene Gemeinschaften', 'Verbraucher und Endnutzer']:
                return 'Social'  # Zuordnung der Kategorie "Sozial"
            elif theme == 'Unternehmenspolitik':
                return 'Governance'  # Zuordnung der Kategorie "Governance"
            else:
                return 'Sonstige'  # Andere Themen

        # Farbe für jedes Thema zuweisen
        selected_columns_df['color'] = selected_columns_df['Thema'].apply(assign_color)

        # Erstellen eines Streudiagramms
        scatter = alt.Chart(selected_columns_df, width=1000, height=800).mark_circle().encode(
            # X-Achse: Finanzielle Wesentlichkeit
            x=alt.X('Score Finanzen', scale=alt.Scale(domain=(0, 1000)), title='Finanzielle Wesentlichkeit'),
            # Y-Achse: Auswirkungsbezogene Wesentlichkeit
            y=alt.Y('Score Auswirkung', scale=alt.Scale(domain=(0, 1000)), title='Auswirkungsbezogene Wesentlichkeit'),
            # Farbe: Zuordnung basierend auf dem Thema (Kategorie)
            color=alt.Color('color:N', scale=alt.Scale(
                domain=['Environmental', 'Social', 'Governance', 'Sonstige'],
                range=['green', 'yellow', 'blue', 'gray']
            ), legend=alt.Legend(
                title="Thema",  # Titel der Legende
                orient="right",  # Position der Legende
                titleColor='black',  # Farbe des Titels
                labelColor='black',  # Farbe der Beschriftungen
                titleFontSize=12,  # Schriftgröße des Titels
                labelFontSize=10,  # Schriftgröße der Beschriftungen
                values=['Environmental', 'Social', 'Governance', 'Sonstige']
            )),
            # Punktgröße: Basierend auf der Wichtigkeit der Stakeholder
            size=alt.Size('Stakeholder Wichtigkeit:Q', scale=alt.Scale(range=[100, 1000]), legend=alt.Legend(
                title="Stakeholder Wichtigkeit",
                orient="right",
                titleColor='black',
                labelColor='black',
                titleFontSize=12,
                labelFontSize=10
            )),
            # Tooltip, das Informationen über die Punkte anzeigt
            tooltip=['Score Finanzen', 'Score Auswirkung', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Stakeholder Wichtigkeit']
        )

        # Hinzufügen einer roten Diagonallinie
        line = alt.Chart(pd.DataFrame({
            'x': [0, st.session_state['intersection_value']],
            'y': [st.session_state['intersection_value'], 0]
        })).mark_line(color='red').encode(
            x='x:Q',
            y='y:Q'
        )

        # Hinzufügen eines schattierten Bereichs unter der Linie
        area = alt.Chart(pd.DataFrame({
            'x': [0, st.session_state['intersection_value']],
            'y': [st.session_state['intersection_value'], 0]
        })).mark_area(opacity=0.3, color='lightcoral').encode(
            x='x:Q',
            y='y:Q'
        )

        # Kombinieren von Streudiagramm, Linie und schattiertem Bereich
        chart = area + scatter + line


        # Speichere das Diagramm als PNG-Datei mit vl-convert
        chart_png_file = os.path.join(os.path.dirname(__file__), '..', 'Pictures', 'scatter_chart.png')
        chart.save(chart_png_file, format='png')

        # Öffne das 'Shortlist'-Arbeitsblatt aus der Excel-Datei
        chart_sheet = workbook['Shortlist']

        # Füge Beschriftungen für den Schwellenwert und die Wesentlichkeitsmatrix ein
        chart_sheet['H1'] = "Schwellenwert für wesentliche Themen:"
        chart_sheet['H2'] = "Schwellenwert für Stakeholder Wichtigkeit:"
        chart_sheet['H3'] = "Wesentlichkeitsmatrix der Shortlist:"

        # Füge das PNG-Bild in das Excel-Arbeitsblatt ein
        img = ExcelImage(chart_png_file)
        img.width, img.height = 600, 400  # Bestimme die gewünschte Größe des Bildes
        chart_sheet.add_image(img, 'H5')  # Platziere das Bild in der Zelle H5

        # Füge die Werte für den Schwellenwert und die Stakeholder-Wichtigkeit ein
        chart_sheet['L1'] = st.session_state['intersection_value']  # Setze den Schnittpunktwert ein
        chart_sheet['L2'] = st.session_state['stakeholder_importance_value']  # Setze den Wert für die Stakeholder-Wichtigkeit ein

    # Überprüfe, ob Daten für das Diagramm vorhanden sind, falls nicht, zeige eine Info-Nachricht an
    else:
        pass

        # ---------- Wesentlichkeitsmatrix ohne Schwellenwerte ---------- #

        # Überprüfung, ob die Daten für das Diagramm im Session-State vorhanden sind
        if 'selected_columns' in st.session_state and len(st.session_state['selected_columns']) > 0:
            selected_columns = st.session_state['selected_columns']
            
            # Daten vorbereiten
            selected_columns_df = pd.DataFrame(selected_columns)
            # Die Spalten, die im Diagramm angezeigt werden sollen
            columns_to_display = ['Score Finanzen', 'Score Auswirkung', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Stakeholder Wichtigkeit']
            
            # Überprüfe, ob die Spalte 'Thema' existiert und gehe mit fehlenden Werten um
            if 'Thema' not in selected_columns_df.columns:
                st.error("Die Spalte 'Thema' fehlt in den Daten.")
                return
            
            # Entferne Zeilen mit fehlenden Werten in der Spalte 'Thema'
            selected_columns_df = selected_columns_df.dropna(subset=['Thema'])
            selected_columns_df['Thema'] = selected_columns_df['Thema'].astype(str)

            # Funktion zur Zuweisung einer Farbe basierend auf dem Thema
            def assign_color(theme):
                if theme in ['Klimawandel', 'Umweltverschmutzung', 'Wasser- & Meeresressourcen', 'Biodiversität', 'Kreislaufwirtschaft']:
                    return 'Environmental'  # Zuordnung zur Kategorie "Umwelt"
                elif theme in ['Eigene Belegschaft', 'Belegschaft Lieferkette', 'Betroffene Gemeinschaften', 'Verbraucher und Endnutzer']:
                    return 'Social'  # Zuordnung zur Kategorie "Sozial"
                elif theme == 'Unternehmenspolitik':
                    return 'Governance'  # Zuordnung zur Kategorie "Governance"
                else:
                    return 'Sonstige'  # Andere Themen

            # Farbe für jedes Thema zuweisen
            selected_columns_df['color'] = selected_columns_df['Thema'].apply(assign_color)

            # Erstelle das Streudiagramm
            scatter = alt.Chart(selected_columns_df, width=1000, height=800).mark_circle().encode(
                # X-Achse: Finanzielle Wesentlichkeit
                x=alt.X('Score Finanzen', scale=alt.Scale(domain=(0, 1000)), title='Finanzielle Wesentlichkeit'),
                # Y-Achse: Auswirkungsbezogene Wesentlichkeit
                y=alt.Y('Score Auswirkung', scale=alt.Scale(domain=(0, 1000)), title='Auswirkungsbezogene Wesentlichkeit'),
                # Farbe: Zuordnung basierend auf dem Thema (Kategorie)
                color=alt.Color('color:N', scale=alt.Scale(
                    domain=['Environmental', 'Social', 'Governance', 'Sonstige'],
                    range=['green', 'yellow', 'blue', 'gray']
                ), legend=alt.Legend(
                    title="Thema",  # Titel der Legende
                    orient="right",  # Position der Legende
                    titleColor='black',  # Farbe des Titels
                    labelColor='black',  # Farbe der Beschriftungen
                    titleFontSize=12,  # Schriftgröße des Titels
                    labelFontSize=10  # Schriftgröße der Beschriftungen
                )),
                # Punktgröße: Basierend auf der Wichtigkeit der Stakeholder
                size=alt.Size('Stakeholder Wichtigkeit:Q', scale=alt.Scale(range=[100, 1000]), legend=alt.Legend(
                    title="Stakeholder Wichtigkeit",
                    orient="right",
                    titleColor='black',
                    labelColor='black',
                    titleFontSize=12,
                    labelFontSize=10
                )),
                # Tooltip für zusätzliche Informationen bei Hover
                tooltip=['Score Finanzen', 'Score Auswirkung', 'Thema', 'Unterthema', 'Unter-Unterthema', 'Stakeholder Wichtigkeit']
            )

            # Speichere das Streudiagramm als PNG-Datei
            chart_png_file_2 = os.path.join(os.path.dirname(__file__), '..', 'Pictures', 'scatter_chart_without_thresholds.png')
            chart.save(chart_png_file_2, format='png')

            # Öffne das 'Übersicht'-Arbeitsblatt aus der Excel-Datei
            chart_sheet_2 = workbook['Übersicht']

            # Füge das zweite PNG-Bild in das Excel-Arbeitsblatt ein
            img = ExcelImage(chart_png_file_2)
            img.width, img.height = 600, 400  # Setze die gewünschte Bildgröße
            chart_sheet_2.add_image(img, 'C19')  # Platziere das Bild in der Zelle C19

        else:
            pass

        # ---------- Fortschritt WA (Wesentlichkeitsanalyse) ---------- #

        # Wähle das 'Übersicht'-Arbeitsblatt aus der Excel-Datei aus
        overview_sheet = workbook['Übersicht']

        # Aktualisiere das Arbeitsblatt ab Zelle C3 mit den Fortschrittsdaten der Wesentlichkeitsanalyse
        session_states_to_check = [
            ('checkbox_state_1', '1. Stakeholder Management'),
            ('checkbox_state_2', '2. Stakeholder Auswahl'),
            ('checkbox_state_3', '3. Themenspezifische ESRS'),
            ('checkbox_state_4', '4. Interne Nachhaltigkeitspunkte'),
            ('checkbox_state_5', '5. Externe Nachhaltigkeitspunkte'),
            ('checkbox_state_6', '6. Bewertung der Longlist'),
            ('checkbox_state_7', '7. Shortlist')
        ]

        # Starte in Zeile 3 für die Aufgaben
        row_start = 3  # Beginn bei C3
        col_name = 'B'  # Spalte für die Aufgabennamen
        col_status = 'C'  # Spalte für den Status

        # Rufe den aktuellen Fortschritt ab und schreibe ihn in die Excel-Datei
        completed_count = 0
        for key, name in session_states_to_check:
            overview_sheet[f'{col_name}{row_start}'] = name
            
            # Sonderbehandlung für checkbox_state_5 (Externe Nachhaltigkeitspunkte)
            if key == 'checkbox_state_5':
                if 'Einbezogene_Stakeholder' in st.session_state and 'sidebar_companies' in st.session_state:
                    if not st.session_state['Einbezogene_Stakeholder'] and not st.session_state['sidebar_companies']:
                        overview_sheet[f'{col_status}{row_start}'] = "✘"
                    elif not st.session_state['Einbezogene_Stakeholder']:
                        overview_sheet[f'{col_status}{row_start}'] = "✔"
                        completed_count += 1
                    else:
                        count = len([opt for opt in st.session_state['Einbezogene_Stakeholder'] if opt not in st.session_state['sidebar_companies']])
                        if st.session_state.get(key) == True:
                            overview_sheet[f'{col_status}{row_start}'] = "✔"
                            completed_count += 1
                        else:
                            overview_sheet[f'{col_status}{row_start}'] = f"Es fehlt noch {count} Stakeholder."
                else:
                    overview_sheet[f'{col_status}{row_start}'] = "✘"
            
            # Sonderbehandlung für checkbox_state_3 (Themenspezifische ESRS)
            elif key == 'checkbox_state_3':
                if st.session_state.get(key) == True:
                    overview_sheet[f'{col_status}{row_start}'] = "✔"
                    completed_count += 1
                else:
                    if 'checkbox_count' in st.session_state and st.session_state['checkbox_count'] > 0:
                        percentage_missing = calculate_percentages()  # Annahme: Diese Funktion existiert
                        overview_sheet[f'{col_status}{row_start}'] = f"Es fehlen noch {percentage_missing}%."
                    else:
                        overview_sheet[f'{col_status}{row_start}'] = "✘"
            
            # Sonderbehandlung für checkbox_state_6 (Bewertung der Longlist)
            elif key == 'checkbox_state_6':
                if 'longlist' in st.session_state and not st.session_state['longlist'].empty:
                    if st.session_state.get(key) == True:
                        overview_sheet[f'{col_status}{row_start}'] = "✔"
                        completed_count += 1
                    else:
                        nein_prozent = bewertung_Uebersicht_Nein()  # Annahme: Diese Funktion existiert
                        overview_sheet[f'{col_status}{row_start}'] = f"Es fehlen noch {nein_prozent}%."
                else:
                    overview_sheet[f'{col_status}{row_start}'] = "✘"
            
            # Standardbehandlung für andere Zustände
            else:
                if st.session_state.get(key) == True:
                    overview_sheet[f'{col_status}{row_start}'] = "✔"
                    completed_count += 1
                else:
                    overview_sheet[f'{col_status}{row_start}'] = "✘"
            
            row_start += 1

        # ---------- Fortschritt Themenspezifische ESRS ---------- #

        # Füge die Anzahl der Checkboxen und den Fortschrittsprozentsatz in die Zelle C11 des Arbeitsblattes 'Übersicht' ein
        if 'checkbox_count' in st.session_state:
            checkbox_count = st.session_state['checkbox_count']
            total_checkboxes = 93  # Gesamtzahl der Checkboxen
            number_of_missing_checkboxes = total_checkboxes - checkbox_count
            
            # Berechne den Prozentsatz der abgeschlossenen Checkboxen
            percentage_complete = round((checkbox_count / total_checkboxes) * 100, 0)
            
            # Schreibe die Anzahl der Checkboxen in C11
            overview_sheet['C11'] = checkbox_count
            
            # Schreibe die verbleibenden Checkboxen und den Prozentsatz in C12
            overview_sheet['C12'] = number_of_missing_checkboxes

        # ---------- Anzahl Shortlist Punkte ---------- #

        # Füge die Anzahl der Shortlist-Punkte in die Zelle C17 des Arbeitsblattes 'Übersicht' ein
        count_shortlist = 0  # Standardwert ist 0
        if 'filtered_df' in st.session_state and not st.session_state['filtered_df'].empty:
            count_shortlist = len(st.session_state['filtered_df'])

        # Schreibe die Anzahl der Shortlist-Punkte in C17
        overview_sheet['C17'] = count_shortlist


    # ---------- Fortschritt Longlist ---------- #

    # Füge die Anzahl der Longlist-Punkte in die Zelle C14 des 'Übersicht'-Blatts ein
    count_longlist = 0  # Standardwert ist 0
    if 'combined_df' in st.session_state and not st.session_state.combined_df.empty:
        count_longlist = len(st.session_state.combined_df)

    # Schreibe die Longlist-Anzahl in die Zelle C14
    overview_sheet['C14'] = count_longlist 

    # ---------- Fortschritt Bewertungen in der Longlist ---------- #

    # Füge die Ja- und Nein-Bewertungen in die Zellen C15 und D15 des 'Übersicht'-Blatts ein
    if 'longlist' in st.session_state and not st.session_state['longlist'].empty:
        # Zähle die Bewertungen in der Longlist
        bewertung_counts = st.session_state['longlist']['Bewertet'].value_counts()

        # Berechne die Anzahl der Ja-Bewertungen
        ja_bewertungen = bewertung_counts.get('Ja', 0)

        # Schreibe die Anzahl der Ja-Bewertungen in die Zelle C15
        overview_sheet['C15'] = ja_bewertungen

    # ---------- Shortlist Sheet ---------- #

    # Überprüfe, ob das 'Shortlist'-Blatt existiert
    if 'Shortlist' not in workbook.sheetnames:
        st.error("Das Blatt 'Shortlist' existiert nicht in der Datei.")
        return

    # Wähle das 'Shortlist'-Blatt aus der Excel-Datei
    shortlist_sheet = workbook['Shortlist']

    # Hole die gefilterten Daten aus dem Session-State
    dataframe = st.session_state.filtered_df

    # Schreibe die gefilterten Daten (Thema, Unterthema, Unter-Unterthema) in das 'Shortlist'-Blatt
    first_empty_row = 1
    while shortlist_sheet[f'A{first_empty_row}'].value or shortlist_sheet[f'B{first_empty_row}'].value or shortlist_sheet[f'C{first_empty_row}'].value:
        first_empty_row += 1

    # Schreibe die gefilterten Daten (Thema, Unterthema, Unter-Unterthema) in das 'Shortlist'-Blatt
    for index, row in dataframe.iterrows():
        shortlist_sheet[f'A{first_empty_row}'] = row['Thema']
        shortlist_sheet[f'B{first_empty_row}'] = row['Unterthema']
        shortlist_sheet[f'C{first_empty_row}'] = row['Unter-Unterthema']
        shortlist_sheet[f'D{first_empty_row}'] = row['Score Finanzen']
        shortlist_sheet[f'E{first_empty_row}'] = row['Score Auswirkung']
        shortlist_sheet[f'F{first_empty_row}'] = row['Stakeholder Wichtigkeit']
        first_empty_row += 1

    # ---------- Allgemeine Angaben Sheet ---------- #

    # Überprüfe, ob 'Antwort' im Session-State existiert
    if 'Antwort' in st.session_state:
        # Überprüfe, ob das 'Allgemeine Angaben'-Blatt existiert
        if 'Allgemeine Angaben' not in workbook.sheetnames:
            st.error("Das Blatt 'Allgemeine Angaben' existiert nicht in der Datei.")
            return

        # Wähle das 'Allgemeine Angaben'-Blatt aus der Excel-Datei
        general_sheet = workbook['Allgemeine Angaben']

        # Schreibe die 'Antwort'-Daten in Spalte C, beginnend ab Zeile 2
        antwort_data = st.session_state.Antwort
        row_index = 2  # Beginne ab Zeile 2

        for entry in antwort_data:
            general_sheet[f'C{row_index}'] = entry
            row_index += 1

    # ---------- Mindestangaben Sheet ---------- #

    # Füge die Mindestangaben in das "Mindestangaben"-Blatt ein, falls es existiert
    if 'Antworten_MDR' in st.session_state:
        if 'Mindestangaben' not in workbook.sheetnames:
            st.error("Das Blatt 'Mindestangaben' existiert nicht in der Datei.")
            return
        
        # Wähle das 'Mindestangaben'-Blatt aus der Excel-Datei
        min_sheet = workbook['Mindestangaben']

        # Schreibe die 'Antworten_MDR'-Daten in Spalte C, beginnend ab Zeile 2
        mdr_data = st.session_state['Antworten_MDR']
        row_index = 2  # Beginne ab Zeile 2 im 'Mindestangaben'-Blatt

        for entry in mdr_data:
            min_sheet[f'C{row_index}'] = entry
            row_index += 1

   # ---------- Stakeholder Ranking ---------- #

    # Überprüfe, ob das 'Stakeholder'-Blatt existiert
    if 'Stakeholder' not in workbook.sheetnames:
        st.error("Das Blatt 'Stakeholder' existiert nicht in der Datei.")
        return

    # Wähle das 'Stakeholder'-Blatt aus der Excel-Datei
    stakeholder_sheet = workbook['Stakeholder']

    # Hole die Ranking-Daten aus dem Session-State
    if 'ranking_table' not in st.session_state:
        st.error("Es gibt keine 'ranking_table' im Session-State.")
        return

    ranking_table = st.session_state['ranking_table']

    # Schreibe die 'ranking_table'-Daten in das 'Stakeholder'-Blatt, beginnend ab Zeile 3
    first_empty_row_stakeholder = 3  # Beginne ab der ersten leeren Zeile

    for index, row in ranking_table.iterrows():
        stakeholder_sheet[f'A{first_empty_row_stakeholder}'] = row['Ranking']
        stakeholder_sheet[f'B{first_empty_row_stakeholder}'] = row['Gruppe']
        stakeholder_sheet[f'C{first_empty_row_stakeholder}'] = row['Score']
        first_empty_row_stakeholder += 1

    # Exportiere die 'table_right_df'-Daten in das 'Stakeholder'-Blatt, beginnend bei Zelle E3
    if 'Einbezogene_Stakeholder' in st.session_state and st.session_state.Einbezogene_Stakeholder:
        # Filtere die 'ranking_table' für die einbezogenen Stakeholder
        table_right_df = st.session_state['ranking_table'][st.session_state['ranking_table']['Gruppe'].isin(st.session_state.Einbezogene_Stakeholder)]

        # Beginne das Schreiben in Spalte E, Zeile 3
        row = 3
        for index, data in table_right_df.iterrows():
            stakeholder_sheet[f'E{row}'] = data['Ranking']
            stakeholder_sheet[f'F{row}'] = data['Gruppe']
            stakeholder_sheet[f'G{row}'] = data['Score']
            row += 1

    # ---------- Externe Nachhaltigkeitspunkte ---------- #

    # Exportiere den 'stakeholder_punkte_filtered'-DataFrame in das 'Externe Nachhaltigkeitspunkte'-Blatt, beginnend ab Zeile 2
    if 'stakeholder_punkte_filtered' in st.session_state:
        # Überprüfe, ob das 'Externe Nachhaltigkeitspunkte'-Blatt existiert
        if 'Externe Nachhaltigkeitspunkte' not in workbook.sheetnames:
            # Erstelle ein neues Blatt, falls es nicht existiert
            workbook.create_sheet('Externe Nachhaltigkeitspunkte')

        # Wähle das 'Externe Nachhaltigkeitspunkte'-Blatt aus der Excel-Datei
        sustainability_sheet = workbook['Externe Nachhaltigkeitspunkte']

        # Schreibe die Spalten des DataFrames in das Arbeitsblatt
        punkt_df = st.session_state['stakeholder_punkte_filtered']
        row_start = 2
        for idx, row in punkt_df.iterrows():
            sustainability_sheet[f'A{row_start}'] = row.get('Platzierung', '')  # Platzierung in Spalte A
            sustainability_sheet[f'B{row_start}'] = row.get('Thema', '')        # Thema in Spalte B
            sustainability_sheet[f'C{row_start}'] = row.get('Unterthema', '')   # Unterthema in Spalte C
            sustainability_sheet[f'D{row_start}'] = row.get('Unter-Unterthema', '') # Unter-Unterthema in Spalte D
            sustainability_sheet[f'E{row_start}'] = row.get('Stakeholder Bew Auswirkung', '')  # Auswirkung
            sustainability_sheet[f'F{row_start}'] = row.get('Stakeholder Bew Finanzen', '')   # Finanzen
            sustainability_sheet[f'G{row_start}'] = row.get('Stakeholder Gesamtbew', '')     # Gesamtbewertung
            sustainability_sheet[f'H{row_start}'] = row.get('Stakeholder', '')               # Stakeholder
            sustainability_sheet[f'I{row_start}'] = row.get('Quelle', '')                    # Quelle
            row_start += 1

    # ---------- Longlist ---------- #

    # Überprüfe, ob das 'Longlist'-Blatt existiert
    if 'Longlist' not in workbook.sheetnames:
        # Erstelle ein neues Blatt, falls es nicht existiert
        workbook.create_sheet('Longlist')

    # Wähle das 'Longlist'-Blatt aus der Excel-Datei
    longlist_sheet = workbook['Longlist']

    # Überprüfe, ob die Longlist im Session-State vorhanden ist
    if 'longlist' in st.session_state:
        # Hole die Longlist aus dem Session-State
        longlist_df = st.session_state['longlist']

        # Schreibe die Daten in das Arbeitsblatt, beginnend ab Zeile 2
        for r_idx, row in longlist_df.iterrows():
            for c_idx, value in enumerate(row, start=1):
                longlist_sheet.cell(row=r_idx + 2, column=c_idx, value=value)  # Schreibe die Daten ab Zeile 2
    else:
        st.info("Keine Daten für 'Longlist' vorhanden.")

    # ---------- Interne Nachhaltigkeitspunkte ---------- #

    # Überprüfe, ob das 'Interne Nachhaltigkeitspunkte'-Blatt existiert
    if 'Interne Nachhaltigkeitspunkte' not in workbook.sheetnames:
        # Erstelle ein neues Blatt, falls es nicht existiert
        workbook.create_sheet('Interne Nachhaltigkeitspunkte')

    # Wähle das 'Interne Nachhaltigkeitspunkte'-Blatt aus der Excel-Datei
    internal_sustainability_sheet = workbook['Interne Nachhaltigkeitspunkte']

    # Überprüfe, ob 'df2' im Session-State vorhanden ist
    if 'df2' in st.session_state:
        # Hole 'df2' aus dem Session-State
        df2 = st.session_state['df2']

        # Schreibe die Daten in das 'Interne Nachhaltigkeitspunkte'-Blatt
        for r_idx, row in df2.iterrows():
            for c_idx, value in enumerate(row, start=1):
                internal_sustainability_sheet.cell(row=r_idx + 2, column=c_idx, value=value)  # Schreibe die Daten ab Zeile 2
    else:
        st.info("Keine Daten für 'Interne Nachhaltigkeitspunkte' vorhanden.")

    # Speichere die neue Excel-Datei mit dem Namen 'Ergebnisse_WA.xlsx'
    output_file = 'Ergebnisse_WA.xlsx'

    try:
        workbook.save(output_file)
        
        # Biete einen Download-Link an
        with open(output_file, 'rb') as file:
            btn = st.download_button(
                label="Ergebnisse herunterladen",
                data=file,
                file_name=output_file,
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
    except Exception as e:
        st.error(f"Fehler beim Speichern der Excel-Datei: {str(e)}")

# Funktion zur Anzeige der Seite
def display_page():
    st.title("Excel-Ausleitung")
    st.write("Klicken Sie auf den Button, um die Daten in eine Excel-Datei zu exportieren.")
    Ausleitung_Excel()
