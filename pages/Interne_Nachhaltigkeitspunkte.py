import streamlit as st
import pandas as pd
import pickle
import os
from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode, DataReturnMode
from openpyxl import load_workbook
import shutil
import io

# Pfade und Konstanten
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), 'Templates', 'Stakeholder_Input_Vorlage_V1.xlsx')
TEMP_EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'Templates', 'Stakeholder_Input_Vorlage_V1_Copy.xlsx')  # Pfad für die temporäre Excel-Kopie
STATE_FILE = 'SessionStates.pkl'  # Datei zum Speichern des Sitzungszustands

# -------------------- Sitzungszustand --------------------

# Funktion zum Speichern des aktuellen Sitzungszustands
def save_state():
    """
    Speichert den aktuellen Zustand der Session in eine Datei.
    Die Datei wird im Binärformat mittels Pickle gespeichert.
    """
    with open(STATE_FILE, 'wb') as f:  # Öffnet die Datei zum Schreiben im Binärmodus
        pickle.dump(dict(st.session_state), f)  # Speichert den Zustand von st.session_state

# Funktion zur Initialisierung des Sitzungszustands
def initialize_session_state():
    """
    Lädt den gespeicherten Zustand der Session aus einer Datei und initialisiert
    Standardwerte für die Session-Variablen.
    """
    if os.path.exists(STATE_FILE) and os.path.getsize(STATE_FILE) > 0:  # Prüft, ob die Datei existiert und nicht leer ist
        with open(STATE_FILE, 'rb') as f:  # Öffnet die Datei zum Lesen im Binärmodus
            loaded_state = pickle.load(f)  # Lädt den gespeicherten Zustand mit Pickle
            for key, value in loaded_state.items():  # Iteriert über die geladenen Schlüssel-Wert-Paare
                if key not in st.session_state:  # Falls der Schlüssel nicht im aktuellen Session-Zustand existiert
                    st.session_state[key] = value  # Setzt den geladenen Zustand in st.session_state

    # Initialisiert einen leeren DataFrame, falls dieser noch nicht vorhanden ist
    if 'df2' not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Thema", "Unterthema", "Unter-Unterthema"])

    # Initialisiert den Zustand der Checkbox, falls dieser noch nicht existiert
    if 'checkbox_state_4' not in st.session_state:
        st.session_state['checkbox_state_4'] = False  # Setzt den Standardwert auf False

# Funktion zur Checkbox für "Abgeschlossen"
def check_abgeschlossen_intern():
    """
    Erstellt eine Checkbox im UI und speichert deren Zustand in der Session.
    """
    if 'checkbox_state_4' not in st.session_state:  # Initialisiert die Checkbox, falls sie nicht vorhanden ist
        st.session_state['checkbox_state_4'] = False  # Setzt den Standardwert auf False

    # Zeigt die Checkbox im UI an und speichert den Zustand in st.session_state
    st.session_state['checkbox_state_4'] = st.checkbox("Abgeschlossen", value=st.session_state['checkbox_state_4'])
    save_state()  # Speichert den aktuellen Zustand nach jeder Änderung

# -------------------- UI Funktionen --------------------

# Formular zum Hinzufügen neuer Einträge in die Tabelle
def add_entry_form():
    """
    Erstellt ein Formular in der Sidebar, das es ermöglicht, 
    neue Themen und Unterthemen hinzuzufügen.
    """
    with st.sidebar:  # Definiert, dass das Formular in der Sidebar angezeigt wird
        st.markdown("---")  # Fügt eine horizontale Trennlinie in der Sidebar hinzu
        st.write("**Inhalte hinzufügen**")  # Titel für den Eingabebereich

        # Auswahlbox für die Themen, basierend auf den verfügbaren Optionen
        thema = st.selectbox('Thema auswählen', options=get_thema_options(), index=0)

        # Auswahlbox für die Unterthemen, die dynamisch basierend auf dem gewählten Thema geladen werden
        unterthema_options = get_unterthema_options(thema)
        unterthema = st.selectbox('Unterthema auswählen', options=unterthema_options, index=0)
        
        # Textfeld für die Eingabe des Unter-Unterthemas
        unter_unterthema = st.text_input('Unter-Unterthema eingeben')

        # Button, der die Daten in die Tabelle einfügt
        if st.button('➕ Hinzufügen'):  # Überprüft, ob der Hinzufügen-Button gedrückt wurde
            add_row(thema, unterthema, unter_unterthema)  # Fügt die Daten der Tabelle hinzu

# Statische Optionen für die Auswahlbox "Thema"
def get_thema_options():
    """
    Gibt eine Liste aller verfügbaren Hauptthemen zurück.
    """
    return [
        'Klimawandel', 'Umweltverschmutzung', 'Wasser- und Meeresressourcen', 
        'Biologische Vielfalt und Ökosysteme', 'Kreislaufwirtschaft', 'Eigene Belegschaft',
        'Arbeitskräfte in der Wertschöpfungskette', 'Betroffene Gemeinschaften',
        'Verbraucher und End-nutzer', 'Unternehmenspolitik'
    ]

# Dynamische Unterthema-Optionen basierend auf dem ausgewählten Thema
def get_unterthema_options(thema):
    """
    Gibt eine Liste von Unterthemen basierend auf dem ausgewählten Hauptthema zurück.
    """
    options = {  # Eine Zuordnung von Themen zu deren Unterthemen
        'Klimawandel': ['Anpassung an den Klimawandel', 'Klimaschutz', 'Energie'],
        'Umweltverschmutzung': ['Luftverschmutzung', 'Wasserverschmutzung', 'Bodenverschmutzung', 'Verschmutzung von lebenden Organismen und Nahrungsressourcen', 'Besorgniserregende Stoffe', 'Mikroplastik'],
        'Wasser- und Meeresressourcen': ['Wasser', 'Meeresressourcen'],
        'Biologische Vielfalt und Ökosysteme': ['Direkte Ursachen des Biodiversitätsverlusts', 'Auswirkungen auf den Zustand der Arten', 'Auswirkungen auf den Umfang und den Zustand von Ökosystemen'],
        'Kreislaufwirtschaft': ['Auswirkungen und Abhängigkeiten von Ökosystemdienstleistungen', 'Ressourcenzuflüsse, einschließlich Ressourcennutzung', 'Ressourcenabflüsse im Zusammenhang mit Produkten und Dienstleistungen', 'Abfälle'],
        'Eigene Belegschaft': ['Arbeitsbedingungen', 'Gleichbehandlung und Chancengleichheit für alle', 'Sonstige arbeitsbezogene Rechte'],
        'Arbeitskräfte in der Wertschöpfungskette': ['Arbeitsbedingungen', 'Gleichbehandlung und Chancengleichheit für alle', 'Sonstige arbeitsbezogene Rechte'],
        'Betroffene Gemeinschaften': ['Wirtschaftliche, soziale und kulturelle Rechte von Gemeinschaften', 'Bürgerrechte und politische Rechte von Gemeinschaften', 'Rechte indigener Völker'],
        'Verbraucher und End-nutzer': ['Informationsbezogene Auswirkungen für Verbraucher und/oder Endnutzer', 'Persönliche Sicherheit von Verbrauchern und/oder Endnutzern', 'Soziale Inklusion von Verbrauchern und/oder Endnutzern'],
        'Unternehmenspolitik': ['Unternehmenskultur', 'Schutz von Hinweisgebern (Whistleblowers)', 'Tierschutz', 'Politisches Engagement und Lobbytätigkeiten', 'Management der Beziehungen zu Lieferanten, einschließlich Zahlungspraktiken', 'Korruption und Bestechung']
    }
    return options.get(thema, [])  # Gibt die Unterthemen basierend auf dem Thema zurück

# Funktion zum Hinzufügen einer neuen Zeile in den DataFrame
def add_row(thema, unterthema, unter_unterthema):
    """
    Fügt eine neue Zeile in den DataFrame df2 hinzu, 
    basierend auf den vom Benutzer eingegebenen Werten.
    """
    if 'df2' not in st.session_state:  # Initialisiert den DataFrame, falls er nicht existiert
        st.session_state.df2 = pd.DataFrame(columns=["Thema", "Unterthema", "Unter-Unterthema"])
        
    # Sucht nach einer leeren Zeile im DataFrame, um sie zu füllen
    empty_row_index = st.session_state.df2[
        (st.session_state.df2["Thema"] == "") &  # Prüft, ob das Feld "Thema" leer ist
        (st.session_state.df2["Unterthema"] == "") &  # Prüft, ob das Feld "Unterthema" leer ist
        (st.session_state.df2["Unter-Unterthema"] == "")  # Prüft, ob das Feld "Unter-Unterthema" leer ist
    ].first_valid_index()  # Findet den Index der ersten gültigen (leeren) Zeile

    if empty_row_index is not None:  # Wenn eine leere Zeile gefunden wurde
        st.session_state.df2.at[empty_row_index, "Thema"] = thema  # Füllt das Thema in die leere Zeile ein
        st.session_state.df2.at[empty_row_index, "Unterthema"] = unterthema  # Füllt das Unterthema in die leere Zeile ein
        st.session_state.df2.at[empty_row_index, "Unter-Unterthema"] = unter_unterthema  # Füllt das Unter-Unterthema ein
    else:  # Wenn keine leere Zeile vorhanden ist
        new_row = {"Thema": thema, "Unterthema": unterthema, "Unter-Unterthema": unter_unterthema}  # Erstellt eine neue Zeile
        st.session_state.df2 = st.session_state.df2._append(new_row, ignore_index=True)  # Fügt die neue Zeile in den DataFrame ein

# -------------------- Datenanzeige und Bearbeitung --------------------

# Funktion zum Anzeigen und Bearbeiten der Tabelle
def display_data_table():
    """
    Zeigt den aktuellen DataFrame in einer interaktiven Tabelle an.
    Ermöglicht dem Benutzer das Bearbeiten, Hinzufügen und Löschen von Zeilen.
    """
    if not st.session_state.df2.empty:  # Prüft, ob der DataFrame nicht leer ist
        grid_options = configure_grid_options(st.session_state.df2)  # Konfiguriert die Optionen für die Tabelle
        grid_response = AgGrid(  # Erstellt eine interaktive Tabelle mit AgGrid
            st.session_state.df2.reset_index(),  # Setzt den Index des DataFrame zurück, damit die Zeilen nummeriert sind
            gridOptions=grid_options,  # Verwendet die konfigurierten Optionen für die Tabelle
            fit_columns_on_grid_load=True,  # Passt die Spaltenbreite an die Daten an
            height=300,  # Höhe der Tabelle in Pixeln
            width='100%',  # Breite der Tabelle in Prozent (nimmt die gesamte Breite ein)
            update_mode=GridUpdateMode.MODEL_CHANGED,  # Aktualisiert die Daten im Modell, wenn sie geändert werden
            allow_unsafe_jscode=True,  # Erlaubt die Verwendung von unsicherem JavaScript-Code
            return_mode=DataReturnMode.__members__['AS_INPUT'],  # Gibt die bearbeiteten Daten als Eingabe zurück
            selection_mode='multiple'  # Ermöglicht die Mehrfachauswahl von Zeilen
        )
    else:
        st.info("Keine Daten vorhanden.")  # Zeigt eine Info-Nachricht an, wenn keine Daten vorhanden sind
    
    # Button zum Hinzufügen einer leeren Zeile
    if st.button('Leere Zeile hinzufügen'):  # Überprüft, ob der Button zum Hinzufügen leerer Zeilen gedrückt wurde
        add_empty_row()  # Fügt eine leere Zeile in den DataFrame ein
    
    # Button zum Löschen der ausgewählten Zeilen
    if st.button('Ausgewählte Zeilen löschen'):  # Überprüft, ob der Button zum Löschen gedrückt wurde
        if 'selected_rows' in grid_response:  # Überprüft, ob Zeilen ausgewählt wurden
            delete_selected_rows(grid_response)  # Löscht die ausgewählten Zeilen
        else:
            st.warning('Bitte wählen Sie zuerst die Zeilen aus, die Sie löschen möchten.')

    # Button zum Speichern der Änderungen in der Tabelle
    if st.button('Änderungen speichern'):  # Überprüft, ob der Speichern-Button gedrückt wurde
        st.session_state.df2 = grid_response['data'].set_index('index')  # Aktualisiert den DataFrame mit den geänderten Daten
        save_state()  # Speichert den aktuellen Zustand der Session
        st.success('Änderungen erfolgreich gespeichert.')  # Zeigt eine Erfolgsmeldung an

# Konfiguriert die Optionen für die AgGrid-Tabelle
def configure_grid_options(dataframe):
    """
    Konfiguriert die Optionen für die interaktive Tabelle.
    Diese Optionen bestimmen das Verhalten und die Anzeige der Tabelle.
    """
    gb = GridOptionsBuilder.from_dataframe(dataframe)  # Erstellt Optionen basierend auf dem DataFrame
    gb.configure_default_column(editable=True, resizable=True, sortable=True, filterable=True)  # Konfiguriert Standardoptionen für Spalten
    gb.configure_grid_options(domLayout='autoHeight', enableRowId=True, rowId='index')  # Legt fest, dass die Höhe automatisch an die Daten angepasst wird
    grid_options = gb.build()  # Erstellt die endgültigen Optionen
    grid_options['columnDefs'] = [{'checkboxSelection': True, 'headerCheckboxSelection': True, 'width': 50}] + grid_options['columnDefs']  # Fügt eine Checkbox-Spalte zur Auswahl hinzu
    return grid_options  # Gibt die konfigurierten Optionen zurück

# Fügt eine leere Zeile in den DataFrame ein
def add_empty_row():
    """
    Fügt eine neue leere Zeile in den DataFrame df2 ein.
    """
    empty_row = {"Thema": "", "Unterthema": "", "Unter-Unterthema": ""}  # Erstellt eine leere Zeile mit leeren Feldern
    st.session_state.df2 = st.session_state.df2._append(empty_row, ignore_index=True)  # Fügt die leere Zeile in den DataFrame ein
    save_state()  # Speichert den aktuellen Zustand der Session
    st.rerun()  # Lädt die Seite neu, um die Änderungen anzuzeigen

# Löscht die ausgewählten Zeilen aus der Tabelle
def delete_selected_rows(grid_response):
    """
    Löscht die ausgewählten Zeilen basierend auf den vom Benutzer in der Tabelle ausgewählten Zeilen.
    """
    selected_rows = grid_response['selected_rows']  # Ruft die vom Benutzer ausgewählten Zeilen ab
    selected_indices = [row['index'] for row in selected_rows]  # Extrahiert die Indizes der ausgewählten Zeilen
    st.session_state.df2 = st.session_state.df2.drop(selected_indices)  # Löscht die Zeilen aus dem DataFrame
    save_state()  # Speichert den aktuellen Zustand der Session
    st.rerun()  # Lädt die Seite neu, um die Änderungen anzuzeigen

# -------------------- Excel-Export und -Download --------------------

# Überträgt die Daten aus dem DataFrame in die Excel-Vorlage
def transfer_data_to_excel(dataframe):
    """
    Überträgt die Inhalte des DataFrame df2 in eine Excel-Vorlage.
    Die Daten werden in eine temporäre Kopie der Excel-Vorlage eingefügt.
    """
    shutil.copyfile(TEMPLATE_PATH, TEMP_EXCEL_PATH)  # Erstellt eine Kopie der Excel-Vorlage
    workbook = load_workbook(TEMP_EXCEL_PATH)  # Lädt die Excel-Kopie
    sheet = workbook['Interne Nachhaltigkeitspunkte']  # Öffnet das Arbeitsblatt in der Excel-Datei
    
    first_empty_row = 2  # Startreihe für die Einträge in der Excel-Datei

    # Iteriert über die Zeilen im DataFrame und fügt die Daten in die Excel-Datei ein
    for index, row in dataframe.iterrows():
        sheet[f'A{first_empty_row}'] = row['Thema']  # Fügt das Thema in die Excel-Spalte A ein
        sheet[f'B{first_empty_row}'] = row['Unterthema']  # Fügt das Unterthema in die Excel-Spalte B ein
        sheet[f'C{first_empty_row}'] = row['Unter-Unterthema']  # Fügt das Unter-Unterthema in die Excel-Spalte C ein
        first_empty_row += 1  # Bewegt sich zur nächsten Zeile

    workbook.save(TEMP_EXCEL_PATH)  # Speichert die Änderungen in der Excel-Datei
    st.success('Inhalte erfolgreich zur Excel-Datei hinzugefügt.')  # Zeigt eine Erfolgsmeldung an

# Download der Excel-Datei als Stream
def download_excel():
    """
    Bereitet die Excel-Datei für den Download vor und gibt die Datei als Bytestream zurück.
    """
    workbook = load_workbook(TEMP_EXCEL_PATH)  # Lädt die temporäre Excel-Datei
    with io.BytesIO() as virtual_workbook:  # Erstellt einen Puffer für den Excel-Inhalt im Speicher
        workbook.save(virtual_workbook)  # Speichert die Excel-Datei in den Puffer
        virtual_workbook.seek(0)  # Setzt den Lesezeiger auf den Anfang des Puffers
        return virtual_workbook.read()  # Gibt den Inhalt des Puffers zurück (die Excel-Datei)

# -------------------- Hauptanzeige --------------------

# Hauptanzeige der Seite
def display_page():
    """
    Hauptfunktion zur Darstellung der Benutzeroberfläche.
    Diese Funktion koordiniert das UI, das Anzeigen und Bearbeiten von Daten sowie den Excel-Export.
    """
    initialize_session_state()  # Initialisiert die Session-Variablen
    
    col1, col2 = st.columns([7, 1])  # Erstellt zwei Spalten zur Layoutsteuerung
    
    with col1:
        st.header("Interne Nachhaltigkeitspunkte")  # Überschrift der Seite
    with col2:
        container = st.container()  # Erzeugt einen Container für die Checkbox
        with container:
            check_abgeschlossen_intern()  # Zeigt die Checkbox für den Status "Abgeschlossen" an

    # Beschreibung des Seiteninhalts
    st.markdown("""
        Hier können Sie unternehmensspezifische Nachhaltigkeitspunkte hinzufügen und verwalten. Nutzen Sie die Dropdown-Menüs und Textfelder in der Sidebar oder tragen Sie Inhalte direkt in die Tabelle ein. Achten Sie darauf, sofern Sie Inhlate direkt in der Tabelle bearbeiten, diese mit Enter zu bestätigen und anschließend den Speicher-Button zu drücken. Aktualisieren Sie anschließend die Excel-Datei, laden Sie sie herunter und leiten Sie diese an Ihre relevanten Stakeholder weiter.
    """)
    
    add_entry_form()  # Zeigt das Formular zum Hinzufügen neuer Einträge an
    display_data_table()  # Zeigt die interaktive Tabelle an
    
    st.sidebar.markdown("---")  # Trennlinie in der Sidebar
    st.sidebar.write("**Excel-Datei für Stakeholderumfrage**")  # Beschreibung der Excel-Funktionalität
    
    # Button zum Aktualisieren der Excel-Datei in der Sidebar
    if st.sidebar.button('🔃 Excel aktualisieren'):
        transfer_data_to_excel(st.session_state.df2)  # Überträgt die Daten in die Excel-Datei
    
    # Download-Button für die Excel-Datei
    if st.sidebar.download_button(
        label="⬇️ Excel-Datei herunterladen",  # Beschriftung des Buttons
        data=download_excel(),  # Ruft die Excel-Datei als Bytestream ab
        file_name="Stakeholder_Input.xlsx",  # Dateiname der heruntergeladenen Datei
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # MIME-Typ der Excel-Datei
    ):
        st.success("Download gestartet!")  # Zeigt eine Erfolgsmeldung nach dem Start des Downloads an

    save_state()  # Speichert den aktuellen Zustand der Session